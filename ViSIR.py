#!/usr/bin/env python
# coding: utf-8

# In[4]:


import torch
import torch.nn               as     nn
import torch.optim            as     optim
import torchvision.transforms as     transforms
from   torch.utils.data       import DataLoader, Dataset
from   PIL                    import Image
import numpy                  as     np
import matplotlib.pyplot      as     plt
from   collections            import OrderedDict
import torch.nn               as     nn
import numpy                  as     np
from   collections            import OrderedDict
import pandas                 as     pd
from   datetime               import datetime
from   skimage.metrics        import structural_similarity   as  ssim
import openpyxl
import torch
import openpyxl


# Define the Patch Embedding layer
class PatchEmbedding(nn.Module):
    def __init__(self, image_size=100, patch_size=16, in_channels=3, embed_dim=256):
        super(PatchEmbedding, self).__init__()
        self.image_size = image_size
        self.patch_size = patch_size
        self.in_channels = in_channels
        self.embed_dim = embed_dim
        self.num_patches = (image_size // patch_size) ** 2
        self.projection = nn.Conv2d(in_channels, embed_dim, kernel_size=patch_size, stride=patch_size)

    def forward(self, x):
        B, C, H, W = x.shape
        assert H == W == self.image_size, f"Input image size ({H}*{W}) does not match model image size ({self.image_size}*{self.image_size})"
        x = self.projection(x).reshape(B, self.num_patches, self.embed_dim)
        return x

# Define the SIREN components
def get_mgrid(sidelen, dim=2):
    tensors = tuple(dim * [torch.linspace(-1, 1, steps=sidelen)])
    mgrid = torch.stack(torch.meshgrid(*tensors), dim=-1)
    mgrid = mgrid.reshape(-1, dim)
    return mgrid

class SineLayer(nn.Module):        
    def __init__(self, in_features, out_features, bias=True, is_first=False, omega_0=10):
        super().__init__()
        self.omega_0 = omega_0
        self.is_first = is_first
        self.in_features = in_features
        self.linear = nn.Linear(in_features, out_features, bias=bias)
        self.init_weights()
    
    def init_weights(self):
        with torch.no_grad():
            if self.is_first:
                self.linear.weight.uniform_(-1 / self.in_features, 1 / self.in_features)      
            else:
                self.linear.weight.uniform_(-np.sqrt(6 / self.in_features) / self.omega_0, np.sqrt(6 / self.in_features) / self.omega_0)
        
    def forward(self, input):
        return torch.sin(self.omega_0 * self.linear(input))
    
    def forward_with_intermediate(self, input):
        intermediate = self.omega_0 * self.linear(input)
        return torch.sin(intermediate), intermediate

class Siren(nn.Module):
    def __init__(self, in_features, hidden_features, hidden_layers, out_features, outermost_linear=False, first_omega_0=10, hidden_omega_0=10):
        super().__init__()
        self.net = []
        self.net.append(SineLayer(in_features, hidden_features, is_first=True, omega_0=first_omega_0))
        for i in range(hidden_layers):
            self.net.append(SineLayer(hidden_features, hidden_features, is_first=False, omega_0=hidden_omega_0))
        if outermost_linear:
            final_linear = nn.Linear(hidden_features, out_features)
            with torch.no_grad():
                final_linear.weight.uniform_(-np.sqrt(6 / hidden_features) / hidden_omega_0, np.sqrt(6 / hidden_features) / hidden_omega_0)
            self.net.append(final_linear)
        else:
            self.net.append(SineLayer(hidden_features, out_features, is_first=False, omega_0=hidden_omega_0))
        self.net = nn.Sequential(*self.net)
    
    def forward(self, coords):
        coords = coords.clone().detach().requires_grad_(True)
        output = self.net(coords)
        return output, coords        

    def forward_with_activations(self, coords, retain_grad=False):
        activations = OrderedDict()
        activation_count = 0
        x = coords.clone().detach().requires_grad_(True)
        activations['input'] = x
        for i, layer in enumerate(self.net):
            if isinstance(layer, SineLayer):
                x, intermed = layer.forward_with_intermediate(x)
                if retain_grad:
                    x.retain_grad()
                    intermed.retain_grad()
                activations['_'.join((str(layer.__class__), "%d" % activation_count))] = intermed
                activation_count += 1
            else: 
                x = layer(x)
                if retain_grad:
                    x.retain_grad()
            activations['_'.join((str(layer.__class__), "%d" % activation_count))] = x
            activation_count += 1
        return activations

# Define the custom TransformerEncoderLayer
class TransformerEncoderLayer(nn.Module):
    def __init__(self, embed_dim, num_heads, hidden_dim, siren_hidden_features, siren_hidden_layers):
        super(TransformerEncoderLayer, self).__init__()
        self.attention = nn.MultiheadAttention(embed_dim, num_heads)
        self.siren = Siren(embed_dim, siren_hidden_features, siren_hidden_layers, embed_dim, outermost_linear=True)
        self.layer_norm1 = nn.LayerNorm(embed_dim)
        self.layer_norm2 = nn.LayerNorm(embed_dim)
        self.dropout = nn.Dropout(0.1)

    def forward(self, x):
        # Self-attention
        x_res = x
        x, _ = self.attention(x, x, x)
        x = self.dropout(x)
        x = self.layer_norm1(x + x_res)
        
        # SIREN instead of Feedforward network
        x_res = x
        x, _ = self.siren(x)
        x = self.dropout(x)
        x = self.layer_norm2(x + x_res)
        
        return x

# Define the Vision Transformer model with SIREN
class VisionTransformerWithSiren(nn.Module):
    def __init__(self, image_size=100, patch_size=16, in_channels=3, embed_dim=256, num_heads=8, num_layers=1, siren_hidden_features=512, siren_hidden_layers=1,omega=10):
        super(VisionTransformerWithSiren, self).__init__()
        self.image_size = image_size
        self.omega = omega
        self.in_channels =  in_channels
        self.patch_size = patch_size
        self.patch_embed = PatchEmbedding(image_size, patch_size, in_channels, embed_dim)
        self.pos_encoding = nn.Parameter(torch.randn(1, self.patch_embed.num_patches + 1, embed_dim))
        self.transformer_encoder_layers = nn.ModuleList([TransformerEncoderLayer(embed_dim, num_heads, embed_dim, siren_hidden_features, siren_hidden_layers) for _ in range(num_layers)])
        self.fc = nn.Linear(self.patch_embed.num_patches * embed_dim, 2)  # Adjust the output size for the SIREN input
        self.siren = Siren(in_features=2, hidden_features=512, hidden_layers=1, out_features=3 * (image_size * 3) ** 2, outermost_linear=True,first_omega_0=omega, hidden_omega_0=omega)

    def forward(self, x):
        x = self.patch_embed(x)
        B, N, _ = x.shape
        x += self.pos_encoding[:, :N]
        for encoder_layer in self.transformer_encoder_layers:
            x = encoder_layer(x)
        x = x.view(B, -1)
        x = self.fc(x)  # Adjusted to match SIREN input size
        x, _ = self.siren(x)
        x = x.view(B, 3, self.image_size * 3, self.image_size * 3)
        return x


# Define the dataset and data loader
class CustomDataset(Dataset):
    def __init__(self, low_res_images, high_res_images, transform=None):
        self.low_res_images = low_res_images
        self.high_res_images = high_res_images
        self.transform = transform

    def __len__(self):
        return len(self.low_res_images)

    def __getitem__(self, idx):
        low_res_img = self.low_res_images[idx]
        high_res_img = self.high_res_images[idx]
        if self.transform:
            low_res_img = self.transform(low_res_img)
            high_res_img = self.transform(high_res_img)
        return low_res_img, high_res_img

# Load and preprocess the data
def preprocess_image(image_path, image_size):
    img = Image.open(image_path).convert('RGB')
    transform = transforms.Compose([
        transforms.Resize((image_size, image_size)),
        transforms.ToTensor(),
    ])
    return transform(img)

def load_data(image_size=240,img_num=1):
    # Load low-resolution images
    low_res_images = [preprocess_image(f"LR/{img_num}.png", image_size)]
    low_res_images = torch.stack(low_res_images)
    
    # Load high-resolution images
    high_res_images = [preprocess_image(f"HR/{img_num}.png", 3*image_size)]
    high_res_images = torch.stack(high_res_images)

    return low_res_images, high_res_images


def calculate_ssim(image1, image2):
    ssim_value = ssim(image1, image2, channel_axis=-1, data_range=1.0)
    return ssim_value

# Training and testing functions
def train(model, dataloader, criterion, optimizer, epochs=1000,img_num=1,freq=10, layer=10):
    model.train()
    # Initialization
    steps_til_summary = 10
    opt_loss = 10
    opt_epoch = 1
    opt_lossPSNR = 0
    opt_SSIM = 0
    PSNRTr = []

    # train the network
    for epoch in range(epochs):
        epoch_loss = 0
        for low_res, high_res in dataloader:
            optimizer.zero_grad()
            outputs         = model(low_res)
            loss            = criterion(outputs, high_res)
            max_pixel_value = torch.max(high_res)
            lossPSNR        = 20 * torch.log10(max_pixel_value / torch.sqrt(loss))
            image1          = outputs[0].detach().cpu().permute(1, 2, 0).numpy()
            image2          = high_res[0].detach().cpu().permute(1, 2, 0).numpy()
            SSIM            = calculate_ssim(image1, image2)
            epoch_loss     += loss.item()
            loss.backward()
            optimizer.step()
            PSNRTr.append(lossPSNR)
            
        if not epoch % steps_til_summary:
            print(f"Image # {img_num+1}, Epoch {epoch}, Loss: {loss}, PSNR Loss: {lossPSNR}")
            print(f"Optimum loss is {opt_loss} and it obtained in {opt_epoch}'s epoch, associated PSNR and SSIM are {opt_lossPSNR} and {opt_SSIM} respectively")

        if loss < opt_loss:
            opt_loss     = loss
            opt_epoch    = epoch
            opt_lossPSNR = lossPSNR
            opt_SSIM     = SSIM
            opt_outputs  = outputs
    
        
    workbook = openpyxl.load_workbook(f"FinalScoresVitSIRENF{freq}H{layer}.xlsx")
    sheet = workbook.active
    sheet.cell(row=img_num+2, column=1, value=f'Image # {img_num}')
    sheet.cell(row=img_num+2, column=2, value=f'{opt_loss}')
    sheet.cell(row=img_num+2, column=3, value=f'{opt_lossPSNR}')
    sheet.cell(row=img_num+2, column=4, value=f'{opt_SSIM}')
    sheet.cell(row=img_num+2, column=5, value=f'{opt_epoch}')
    
    workbook.save(f"FinalScoresVitSIRENF{freq}H{layer}.xlsx")
    

    comparison = plt.figure(figsize=(12, 6))
    plt.subplot(1, 2, 1)
    plt.imshow(high_res[0].permute(1, 2, 0))
    plt.title('High Resolution Image')
    plt.subplot(1, 2, 2)
    plt.imshow(opt_outputs[0].detach().cpu().permute(1, 2, 0).numpy())
    plt.title('Model Output')
    plt.savefig(f"ViTSIREN2comp_image#{img_num}_F10H1.png")  # Save the figure
    plt.show()
    
def test(model, dataloader, criterion):
    model.eval()
    test_loss = 0
    with torch.no_grad():
        for low_res, high_res in dataloader:
            outputs = model(low_res)
            loss = criterion(outputs, high_res)
            test_loss += loss.item()
    print(f"Test Loss: {test_loss/len(dataloader)}")


#############################################################
#                 Main Training Section                     #
#############################################################
num_of_images = 1
steps_til_summary = 10
Freq = 20
Layer = 3
EPOCHS = 1000

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1, value='Image Number')
sheet.cell(row=1, column=2, value='Scores')
sheet.cell(row=1, column=3, value='PSNR')
sheet.cell(row=1, column=4, value='SSIM')
sheet.cell(row=1, column=5, value='obtained at epoch number')
sheet.cell(row=1, column=6, value='Elapsed Time')

workbook.save(f"FinalScoresVitSIRENF{Freq}H{Layer}.xlsx")


start_time = datetime.now()
for k in range(num_of_images):    # Load data
    low_res_images, high_res_images = load_data(image_size=100, img_num=k)
    
    dataset = CustomDataset(low_res_images, high_res_images)
    dataloader = DataLoader(dataset, batch_size=4, shuffle=True)

    # Initialize model, criterion, and optimizer
    model = VisionTransformerWithSiren(image_size=100, patch_size=16, in_channels=3, siren_hidden_layers=1, embed_dim=2, num_heads=2, num_layers=1,omega=10)
    criterion = nn.MSELoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001)

    # Training step
    train(model, dataloader, criterion, optimizer, epochs=EPOCHS,img_num=k, freq=Freq, layer=Layer)


end_time = datetime.now()
# Calculate the difference
elapsed_time = end_time - start_time
workbook = openpyxl.load_workbook(f"FinalScoresVitSIRENF{Freq}H{Layer}.xlsx")
sheet = workbook.active
sheet.cell(row=2, column=6, value=f'{elapsed_time}')
workbook.save(f"FinalScoresVitSIRENF{Freq}H{Layer}.xlsx")




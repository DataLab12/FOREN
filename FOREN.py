#!/usr/bin/env python
# coding: utf-8

import torch
import torch.nn as nn
import torch.optim as optim
import torchvision.transforms as transforms
from torch.utils.data import DataLoader, Dataset
from PIL import Image
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from collections import OrderedDict
import torch
import torch.nn as nn
import torch.fft
import numpy as np

def get_mgrid(sidelen, dim=2):
    '''Generates a flattened grid of (x,y,...) coordinates in a range of -1 to 1.
    sidelen: int
    dim: int'''
    tensors = tuple(dim * [torch.linspace(-1, 1, steps=sidelen)])
    mgrid = torch.stack(torch.meshgrid(*tensors), dim=-1)
    mgrid = mgrid.reshape(-1, dim)
    return mgrid

class HighPassFilter(nn.Module):
    def __init__(self, cutoff_frequency):
        super(HighPassFilter, self).__init__()
        self.cutoff_frequency = cutoff_frequency

    def forward(self, x):
        # Compute the FFT of the input
        x_fft = torch.fft.fft2(x)
        
        # Create a high-pass filter mask
        freq_x = torch.fft.fftfreq(x.size(-2))
        freq_y = torch.fft.fftfreq(x.size(-1))
        high_pass_mask = (torch.abs(freq_x[:, None]) + torch.abs(freq_y[None, :]) > self.cutoff_frequency).float()
        
        # Apply the high-pass filter mask
        x_fft_filtered = x_fft * high_pass_mask
        
        # Compute the inverse FFT to get the filtered output
        x_filtered = torch.fft.ifft2(x_fft_filtered).real
        
        return x_filtered

class SIRENHighPass(nn.Module):
    def __init__(self, in_features, out_features, hidden_features, hidden_layers, cutoff_frequency):
        super(SIRENHighPass, self).__init__()
        self.net = []
        self.net.append(nn.Linear(in_features, hidden_features))
        self.net.append(HighPassFilter(cutoff_frequency))
        
        
        for _ in range(hidden_layers):
            self.net.append(nn.Linear(hidden_features, hidden_features))
            self.net.append(HighPassFilter(cutoff_frequency))
        
        self.net.append(nn.Linear(hidden_features, out_features))
        
        self.net = nn.Sequential(*self.net)
    def forward(self, x):
        return self.net(x)


# model = SIRENHighPass(in_features=2, out_features=1, hidden_features=256, hidden_layers=3, cutoff_frequency=0.1)

for low_res, high_res in dataloader:
    net = []
    print(low_res.shape)
    net.append(nn.Linear(100,256))
    net.append(HighPassFilter(0.1))
    
   
    net.append(nn.Linear(256,1))
    net.append(HighPassFilter(0.1))

        
    net = nn.Sequential(*net)
    
    net(low_res)


class CustomDataset(Dataset):
    def __init__(self, low_res_images, high_res_images, transform=None):
        self.low_res_images = low_res_images
        self.high_res_images = high_res_images
        self.transform = transform

    def __len__(self):
        return len(self.low_res_images)

    def __getitem__(self, idx):
        low_res_img = self.low_res_images[idx]
        high_res_img = self.high_res_images[idx]
        if self.transform:
            low_res_img = self.transform(low_res_img)
            high_res_img = self.transform(high_res_img)
        return low_res_img, high_res_img

def preprocess_image(image_path, image_size):
    img = Image.open(image_path).convert('RGB')
    transform = transforms.Compose([
        transforms.Resize((image_size, image_size)),
        transforms.ToTensor(),
    ])
    return transform(img)

def load_data(image_size=240,img_num=20):
    # Load low-resolution images
    low_res_images = [preprocess_image(f"C:/Users/Zigorat/Desktop/TXU/DownScaling/MultiLR/{i}.png", image_size) for i in range(20)]
    low_res_images = torch.stack(low_res_images)
    
    # Load high-resolution images
    high_res_images = [preprocess_image(f"C:/Users/Zigorat/Desktop/TXU/DownScaling/MultiHR/{i}.png", 3*image_size) for i in range(20)]
    high_res_images = torch.stack(high_res_images)

    return low_res_images, high_res_images



# Training and testing functions
def train(model, dataloader, criterion, optimizer, epochs=10,img_num=20):
    model.train()
    steps_til_summary = 10
    opt_loss = 10
    opt_epoch = 1
    opt_lossPSNR = 0
    for epoch in range(epochs):
        epoch_loss = 0
        for low_res, high_res in dataloader:
            optimizer.zero_grad()
            outputs = model(low_res)
            print(outputs.shape)
            print(high_res.shape)
            
            loss = criterion(outputs, high_res)
            max_pixel_value = torch.max(high_res)
            lossPSNR = 20 * torch.log10(max_pixel_value / torch.sqrt(loss))
            loss.backward()
            optimizer.step()
            epoch_loss += loss.item()
            
        if not epoch % steps_til_summary:
            print(f"Image # {img_num+1}, Epoch {epoch}, Loss: {loss}, PSNR Loss: {lossPSNR}")
            print(f"Optimum loss is {opt_loss} and it obtained in {opt_epoch}'s epoch, and associated PSNR loss is {opt_lossPSNR}")

        if loss < opt_loss:
            opt_loss = loss
            opt_epoch = epoch
            opt_lossPSNR = lossPSNR
            opt_outputs = outputs


    workbook = openpyxl.load_workbook(f"FinalScoresVit.xlsx")
    sheet = workbook.active
    sheet.cell(row=img_num+2, column=4, value=f'obtained at epoch # {opt_epoch}')
    sheet.cell(row=img_num+2, column=3, value=f'Total PSNR is {opt_lossPSNR}')
    sheet.cell(row=img_num+2, column=2, value=f'Total Loss is {opt_loss}')
    sheet.cell(row=img_num+2, column=1, value=f'Image # {img_num}')
    
    workbook.save(f"FinalScoresVit.xlsx")
    

    comparison = plt.figure(figsize=(12, 6))
    plt.subplot(1, 2, 1)
    plt.imshow(high_res[0].permute(1, 2, 0))
    plt.title('High Resolution Image')
    plt.subplot(1, 2, 2)
    plt.imshow(opt_outputs[0].detach().cpu().permute(1, 2, 0).numpy())
    plt.title('Model Output')
    plt.savefig(f"comparison_{img_num+1}.png")  # Save the figure
    plt.show()
    
def test(model, dataloader, criterion):
    model.eval()
    test_loss = 0
    with torch.no_grad():
        for low_res, high_res in dataloader:
            outputs = model(low_res)
            loss = criterion(outputs, high_res)
            test_loss += loss.item()
    print(f"Test Loss: {test_loss/len(dataloader)}")


num_of_images = 1
steps_til_summary = 10


workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=4, value='Epoch Number')
sheet.cell(row=1, column=3, value='PSNR')
sheet.cell(row=1, column=2, value='Scores')
sheet.cell(row=1, column=1, value='Image Number')
workbook.save(f"FinalScoresVit.xlsx")

for k in range(num_of_images):    # Load data
    low_res_images, high_res_images = load_data(image_size=100, img_num=20)
    
    # Define dataset and data loader
    dataset = CustomDataset(low_res_images, high_res_images)
    dataloader = DataLoader(dataset, batch_size=4, shuffle=True)

    # Initialize model, criterion, and optimizer
    model = SIRENHighPass(in_features=100, out_features=1, hidden_features=256, hidden_layers=3, cutoff_frequency=0.1)
    criterion = nn.MSELoss()
    optimizer = optim.Adam(model.parameters(), lr=0.001)

    # Train and test the model
    train(model, dataloader, criterion, optimizer, epochs=10,img_num=k)
    # test(model, dataloader, criterion)






